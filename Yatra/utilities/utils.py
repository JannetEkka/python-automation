import inspect
import softest
import logging
from openpyxl import load_workbook, Workbook

class Utils(softest.TestCase):
    def assertListItemText(self, list, value):
        for stop in list:
            print('The text is:',stop.text)
            self.soft_assert(self.assertEqual, stop.text, value)
            if stop.text==value:
                print('Test Passed')
            else:
                print('Test Failed')

        self.assert_all()

    def customLogger(loglevel=logging.DEBUG):
        logger_name = inspect.stack()[1][3]
        logger = logging.getLogger(logger_name)
        logger.setLevel(loglevel)
        fh = logging.FileHandler('automation.log', mode='w')
        formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(name)s : %(message)s',
                                    datefmt=('%d/%m/%Y %I:%M:%S %p'))
        fh.setFormatter(formatter)
        logger.addHandler(fh)
        return logger

    def read_data_from_excel(file_name, sheet):
        datalist = []
        
        wb = load_workbook(filename=file_name)
        sh = wb[sheet]
        row_ct = sh.max_row
        col_ct = sh.max_column

        if row_ct < 2 or col_ct < 1:
                return datalist
        
        for i in range(2, row_ct + 1):
            row = []
            for j in range(1, col_ct + 1):
                row.append(sh.cell(row=i, column=j).value)
            datalist.append(row)
        return datalist
    # def read_data_from_csv():
    #     return datalist1